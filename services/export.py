"""Export test cases to Excel (.xlsx) and Markdown formats."""

import logging
import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger(__name__)

COLUMNS = [
    ("ID", 10),
    ("Module", 18),
    ("Test Case Name", 35),
    ("Priority", 10),
    ("Preconditions", 30),
    ("Steps", 45),
    ("Expected Result", 35),
    ("Type", 15),
    ("Method", 12),
]

PRIORITY_COLORS = {
    "P0": "FF4444",
    "P1": "FF8C00",
    "P2": "4488FF",
    "P3": "888888",
}


def export_excel(test_cases: list[dict], output_path: str) -> str:
    """Export test cases to an Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Write headers
    for col_idx, (name, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = width

    # Write data
    cell_alignment = Alignment(vertical="top", wrap_text=True)
    fields = ["id", "module", "name", "priority", "preconditions", "steps", "expected_result", "type", "method"]

    for row_idx, tc in enumerate(test_cases, 2):
        for col_idx, field in enumerate(fields, 1):
            value = tc.get(field, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border

            # Color-code priority
            if field == "priority":
                color = PRIORITY_COLORS.get(value.upper(), "888888")
                cell.font = Font(bold=True, color=color)
                cell.alignment = Alignment(horizontal="center", vertical="top")

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)
    logger.info("Exported %d test cases to Excel: %s", len(test_cases), output_path)
    return output_path


def export_markdown(test_cases: list[dict], output_path: str) -> str:
    """Export test cases to a Markdown file."""
    lines = ["# Test Cases\n"]
    lines.append(f"**Total: {len(test_cases)} test cases**\n")

    # Summary by priority
    priority_counts = {}
    for tc in test_cases:
        p = tc.get("priority", "P2")
        priority_counts[p] = priority_counts.get(p, 0) + 1
    summary_parts = [f"{k}: {v}" for k, v in sorted(priority_counts.items())]
    lines.append(f"Priority breakdown: {' | '.join(summary_parts)}\n")

    # Table
    lines.append("| ID | Module | Test Case Name | Priority | Preconditions | Steps | Expected Result | Type | Method |")
    lines.append("|---|---|---|---|---|---|---|---|---|")

    for tc in test_cases:
        row = [
            tc.get("id", ""),
            tc.get("module", ""),
            tc.get("name", ""),
            tc.get("priority", ""),
            _escape_md(tc.get("preconditions", "")),
            _escape_md(tc.get("steps", "")),
            _escape_md(tc.get("expected_result", "")),
            tc.get("type", ""),
            tc.get("method", ""),
        ]
        lines.append("| " + " | ".join(row) + " |")

    lines.append("")

    # Detailed view
    lines.append("---\n")
    lines.append("## Detailed Test Cases\n")

    for tc in test_cases:
        lines.append(f"### {tc.get('id', '')} — {tc.get('name', '')}\n")
        lines.append(f"- **Module:** {tc.get('module', '')}")
        lines.append(f"- **Priority:** {tc.get('priority', '')}")
        lines.append(f"- **Type:** {tc.get('type', '')}")
        lines.append(f"- **Preconditions:** {tc.get('preconditions', '')}")
        lines.append(f"- **Steps:**\n")
        # Try to split steps into numbered list
        steps = tc.get("steps", "")
        step_lines = re.split(r"(?:\d+\.\s*|\n)", steps) if steps else []
        step_lines = [s.strip() for s in step_lines if s.strip()]
        if step_lines:
            for i, s in enumerate(step_lines, 1):
                lines.append(f"  {i}. {s}")
        else:
            lines.append(f"  {steps}")
        lines.append(f"\n- **Expected Result:** {tc.get('expected_result', '')}\n")

    content = "\n".join(lines)

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(content)

    logger.info("Exported %d test cases to Markdown: %s", len(test_cases), output_path)
    return output_path


import re


def _escape_md(text: str) -> str:
    """Escape pipe characters and newlines for markdown table cells."""
    if not text:
        return ""
    return text.replace("|", "\\|").replace("\n", " ")
